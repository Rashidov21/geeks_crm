from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView, View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q, Count, Avg
from django.http import JsonResponse, HttpResponse
from django.urls import reverse_lazy, reverse
from datetime import timedelta
import json
import csv
from io import BytesIO

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .models import (
    Lead, LeadStatus, LeadHistory, FollowUp, TrialLesson, SalesProfile,
    WorkSchedule, Leave, SalesKPI, DailyKPI, SalesMessage, SalesMessageRead,
    Offer, Reactivation
)
from accounts.models import User, Branch
from accounts.mixins import RoleRequiredMixin, AdminRequiredMixin, TailwindFormMixin
from courses.models import Course, Group, Room


# ==================== LEAD VIEWS ====================

class LeadKanbanView(LoginRequiredMixin, TemplateView):
    """
    Kanban board - statuslar bo'yicha lidlar
    """
    template_name = 'crm/kanban.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Barcha statuslar
        statuses = LeadStatus.objects.filter(is_active=True).order_by('order')
        
        # Lidlarni olish
        leads = Lead.objects.select_related('status', 'assigned_sales', 'interested_course', 'branch')
        
        # Role bo'yicha filtrlash
        if self.request.user.is_sales:
            leads = leads.filter(assigned_sales=self.request.user)
        elif self.request.user.is_sales_manager:
            if hasattr(self.request.user, 'sales_profile') and self.request.user.sales_profile:
                branch = self.request.user.sales_profile.branch
                if branch:
                    leads = leads.filter(branch=branch)
        
        # Har bir status uchun lidlarni guruhlash
        kanban_data = []
        for status in statuses:
            status_leads = leads.filter(status=status).order_by('-created_at')[:20]
            kanban_data.append({
                'status': status,
                'leads': status_leads,
                'count': leads.filter(status=status).count()
            })
        
        context['kanban_data'] = kanban_data
        context['statuses'] = statuses
        context['total_leads'] = leads.count()
        
        # Stats for Admin/Manager
        if self.request.user.is_admin or self.request.user.is_manager:
            from django.utils import timezone
            today = timezone.now().date()
            context['stats'] = {
                'new_leads': leads.filter(status__code='new').count(),
                'today_followups': FollowUp.objects.filter(
                    due_date__date=today, completed=False
                ).count(),
                'overdue': FollowUp.objects.filter(
                    due_date__lt=timezone.now(), completed=False
                ).count(),
                'trials': TrialLesson.objects.filter(
                    date__gte=today, result__isnull=True
                ).count(),
            }
        
        # Filter options
        context['sales_users'] = User.objects.filter(role__in=['sales', 'sales_manager'])
        context['courses'] = Course.objects.filter(is_active=True)
        context['sources'] = Lead.SOURCE_CHOICES
        
        return context


class LeadTableView(LoginRequiredMixin, ListView):
    """
    Jadval ko'rinishi + filter + export
    """
    model = Lead
    template_name = 'crm/lead_table.html'
    context_object_name = 'leads'
    paginate_by = 50
    
    def get_queryset(self):
        queryset = Lead.objects.select_related('status', 'assigned_sales', 'interested_course', 'branch')
        
        # Role bo'yicha filtrlash
        if self.request.user.is_sales:
            queryset = queryset.filter(assigned_sales=self.request.user)
        elif self.request.user.is_sales_manager:
            if hasattr(self.request.user, 'sales_profile') and self.request.user.sales_profile:
                branch = self.request.user.sales_profile.branch
                if branch:
                    queryset = queryset.filter(branch=branch)
        
        # Filterlar
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status__code=status)
        
        source = self.request.GET.get('source')
        if source:
            queryset = queryset.filter(source=source)
        
        course = self.request.GET.get('course')
        if course:
            queryset = queryset.filter(interested_course_id=course)
        
        sales = self.request.GET.get('sales')
        if sales and (self.request.user.is_admin or self.request.user.is_manager or self.request.user.is_sales_manager):
            queryset = queryset.filter(assigned_sales_id=sales)
        
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) | Q(phone__icontains=search)
            )
        
        return queryset.order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['statuses'] = LeadStatus.objects.filter(is_active=True)
        context['sources'] = Lead.SOURCE_CHOICES
        context['courses'] = Course.objects.filter(is_active=True)
        context['sales_users'] = User.objects.filter(role__in=['sales', 'sales_manager'], is_active=True)
        return context


class LeadExportView(LoginRequiredMixin, View):
    """
    Lidlarni Excel yoki CSV formatida export qilish
    """
    def get(self, request):
        format_type = request.GET.get('format', 'excel')
        
        # Lidlarni olish - LeadTableView bilan bir xil filterlar
        leads = Lead.objects.select_related('status', 'assigned_sales', 'interested_course', 'branch')
        
        # Role bo'yicha filtrlash
        if request.user.is_sales:
            leads = leads.filter(assigned_sales=request.user)
        elif request.user.is_sales_manager:
            if hasattr(request.user, 'sales_profile') and request.user.sales_profile:
                branch = request.user.sales_profile.branch
                if branch:
                    leads = leads.filter(branch=branch)
        
        # Filterlar - LeadTableView bilan bir xil
        status = request.GET.get('status')
        if status:
            leads = leads.filter(status__code=status)
        
        source = request.GET.get('source')
        if source:
            leads = leads.filter(source=source)
        
        course = request.GET.get('course')
        if course:
            leads = leads.filter(interested_course_id=course)
        
        sales = request.GET.get('sales')
        if sales and (request.user.is_admin or request.user.is_manager or request.user.is_sales_manager):
            leads = leads.filter(assigned_sales_id=sales)
        
        search = request.GET.get('search')
        if search:
            from django.db.models import Q
            leads = leads.filter(
                Q(name__icontains=search) | Q(phone__icontains=search)
            )
        
        leads = leads.order_by('-created_at')
        
        if format_type == 'csv':
            return self.export_csv(leads)
        else:
            return self.export_excel(leads)
    
    def export_csv(self, leads):
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="leads.csv"'
        response.write('\ufeff')  # BOM for Excel
        
        writer = csv.writer(response)
        writer.writerow(['Ism', 'Telefon', 'Manba', 'Status', 'Kurs', 'Sotuvchi', 'Filial', 'Sana'])
        
        for lead in leads:
            writer.writerow([
                lead.name,
                lead.phone,
                lead.get_source_display() if lead.source else '',
                lead.status.name if lead.status else '',
                lead.interested_course.name if lead.interested_course else '',
                lead.assigned_sales.get_full_name() if lead.assigned_sales else '',
                lead.branch.name if lead.branch else '',
                lead.created_at.strftime('%d.%m.%Y %H:%M'),
            ])
        
        return response
    
    def export_excel(self, leads):
        if not OPENPYXL_AVAILABLE:
            return self.export_csv(leads)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lidlar"
        
        # Header
        headers = ['Ism', 'Telefon', 'Qo\'shimcha telefon', 'Manba', 'Status', 'Kurs', 'Sotuvchi', 'Filial', 'Sana']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
        
        # Data
        for row, lead in enumerate(leads, 2):
            ws.cell(row=row, column=1, value=lead.name)
            ws.cell(row=row, column=2, value=lead.phone)
            ws.cell(row=row, column=3, value=lead.secondary_phone or '')
            ws.cell(row=row, column=4, value=lead.get_source_display() if lead.source else '')
            ws.cell(row=row, column=5, value=lead.status.name if lead.status else '')
            ws.cell(row=row, column=6, value=lead.interested_course.name if lead.interested_course else '')
            ws.cell(row=row, column=7, value=lead.assigned_sales.get_full_name() if lead.assigned_sales else '')
            ws.cell(row=row, column=8, value=lead.branch.name if lead.branch else '')
            ws.cell(row=row, column=9, value=lead.created_at.strftime('%d.%m.%Y %H:%M'))
        
        # Column width
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="leads.xlsx"'
        return response


class LeadListView(LoginRequiredMixin, ListView):
    """
    Leadlar ro'yxati (Kanban board uchun)
    """
    model = Lead
    template_name = 'crm/lead_list.html'
    context_object_name = 'leads'
    paginate_by = 50
    
    def get_queryset(self):
        queryset = Lead.objects.select_related('status', 'assigned_sales', 'interested_course', 'branch')
        
        if self.request.user.is_sales:
            queryset = queryset.filter(assigned_sales=self.request.user)
        elif self.request.user.is_sales_manager:
            if hasattr(self.request.user, 'sales_profile') and self.request.user.sales_profile:
                branch = self.request.user.sales_profile.branch
                if branch:
                    queryset = queryset.filter(branch=branch)
        
        status_filter = self.request.GET.get('status')
        if status_filter:
            queryset = queryset.filter(status__code=status_filter)
        
        return queryset.order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['statuses'] = LeadStatus.objects.filter(is_active=True).order_by('order')
        total_leads = Lead.objects.count()
        enrolled_leads = Lead.objects.filter(status__code='enrolled').count()
        pending_leads = Lead.objects.filter(
            status__code__in=['new', 'contacted', 'interested']
        ).count()
        overdue_followups = FollowUp.objects.filter(
            completed=False,
            due_date__lt=timezone.now()
        ).count()
        
        # Stats for cards
        context['stats'] = [
            {'label': 'Jami lidlar', 'value': total_leads, 'icon': 'fas fa-users', 'color': 'text-purple-600'},
            {'label': 'Kursga yozilgan', 'value': enrolled_leads, 'icon': 'fas fa-check-circle', 'color': 'text-green-600'},
            {'label': 'Kutilayotgan', 'value': pending_leads, 'icon': 'fas fa-clock', 'color': 'text-yellow-600'},
            {'label': 'Kechikkan follow-up', 'value': overdue_followups, 'icon': 'fas fa-exclamation-triangle', 'color': 'text-red-600'},
        ]
        
        # Permissions
        user = self.request.user
        context['can_create'] = user.is_superuser or (hasattr(user, 'is_admin') and user.is_admin) or (hasattr(user, 'is_manager') and user.is_manager) or (hasattr(user, 'is_sales_manager') and user.is_sales_manager) or (hasattr(user, 'is_sales') and user.is_sales)
        context['can_edit'] = context['can_create']
        context['can_delete'] = user.is_superuser or (hasattr(user, 'is_admin') and user.is_admin) or (hasattr(user, 'is_manager') and user.is_manager)
        return context


class LeadDetailView(LoginRequiredMixin, DetailView):
    """
    Lead batafsil ma'lumotlari
    """
    model = Lead
    template_name = 'crm/lead_detail.html'
    context_object_name = 'lead'
    
    def get_queryset(self):
        queryset = Lead.objects.select_related(
            'status', 'assigned_sales', 'interested_course', 'branch', 
            'trial_group', 'trial_room', 'enrolled_group'
        )
        
        if self.request.user.is_sales:
            queryset = queryset.filter(assigned_sales=self.request.user)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        lead = self.object
        
        # Follow-ups (oxirgi 10 tasi)
        context['followups'] = FollowUp.objects.filter(lead=lead).order_by('-due_date')[:10]
        context['followups_count'] = FollowUp.objects.filter(lead=lead).count()
        
        # History (oxirgi 20 tasi)
        context['history'] = lead.history.all().order_by('-created_at')[:20]
        context['history_count'] = lead.history.count()
        
        # Trial lessons (oxirgi 10 tasi)
        context['trial_lessons'] = TrialLesson.objects.filter(lead=lead).order_by('-date')[:10]
        context['trial_lessons_count'] = TrialLesson.objects.filter(lead=lead).count()
        
        # Faol takliflar
        today = timezone.now().date()
        context['active_offers'] = Offer.objects.filter(
            is_active=True,
            valid_from__lte=today,
            valid_until__gte=today
        ).filter(
            Q(course__isnull=True) | Q(course=lead.interested_course)
        )
        
        context['statuses'] = LeadStatus.objects.filter(is_active=True)
        context['groups'] = Group.objects.filter(is_active=True)
        
        # Permissions
        user = self.request.user
        context['can_edit'] = user.is_superuser or (hasattr(user, 'is_admin') and user.is_admin) or (hasattr(user, 'is_manager') and user.is_manager) or (hasattr(user, 'is_sales_manager') and user.is_sales_manager) or (hasattr(user, 'is_sales') and user.is_sales)
        context['can_delete'] = user.is_superuser or (hasattr(user, 'is_admin') and user.is_admin) or (hasattr(user, 'is_manager') and user.is_manager)
        
        return context


class LeadCreateView(TailwindFormMixin, RoleRequiredMixin, CreateView):
    """
    Yangi lead yaratish
    """
    model = Lead
    template_name = 'crm/lead_form.html'
    fields = ['name', 'phone', 'secondary_phone', 'interested_course', 'source', 'branch', 'notes']
    allowed_roles = ['admin', 'manager', 'sales_manager', 'sales']
    success_url = reverse_lazy('crm:lead_list')
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        
        # Yangi status
        new_status = LeadStatus.objects.filter(code='new').first()
        if new_status:
            form.instance.status = new_status
        
        # Agar sotuvchi yaratsa, o'ziga biriktirish
        if self.request.user.is_sales:
            form.instance.assigned_sales = self.request.user
            form.instance.assigned_at = timezone.now()
        
        messages.success(self.request, 'Lid muvaffaqiyatli yaratildi.')
        return super().form_valid(form)


class LeadUpdateView(TailwindFormMixin, RoleRequiredMixin, UpdateView):
    """
    Leadni tahrirlash
    """
    model = Lead
    template_name = 'crm/lead_form.html'
    fields = ['name', 'phone', 'secondary_phone', 'interested_course', 'source', 
              'status', 'branch', 'notes']
    allowed_roles = ['admin', 'manager', 'sales_manager', 'sales']
    
    def get_queryset(self):
        queryset = Lead.objects.all()
        if self.request.user.is_sales:
            queryset = queryset.filter(assigned_sales=self.request.user)
        return queryset
    
    def get_success_url(self):
        return reverse('crm:lead_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        # Status o'zgarishini tekshirish
        old_lead = Lead.objects.get(pk=self.object.pk)
        if old_lead.status != form.instance.status:
            LeadHistory.objects.create(
                lead=form.instance,
                old_status=old_lead.status,
                new_status=form.instance.status,
                changed_by=self.request.user,
                notes=f"Status o'zgardi"
            )
            
            # Lost status bo'lsa
            if form.instance.status and form.instance.status.code == 'lost':
                form.instance.lost_at = timezone.now()
            
            # Enrolled status bo'lsa
            if form.instance.status and form.instance.status.code == 'enrolled':
                form.instance.enrolled_at = timezone.now()
        
        messages.success(self.request, 'Lid muvaffaqiyatli yangilandi.')
        return super().form_valid(form)


class LeadDeleteView(RoleRequiredMixin, DeleteView):
    """
    Leadni o'chirish
    """
    model = Lead
    template_name = 'crm/lead_confirm_delete.html'
    success_url = reverse_lazy('crm:lead_list')
    allowed_roles = ['admin', 'manager']
    
    def get_queryset(self):
        queryset = Lead.objects.all()
        if self.request.user.is_sales:
            queryset = queryset.filter(assigned_sales=self.request.user)
        return queryset
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Lid o\'chirildi.')
        return super().delete(request, *args, **kwargs)


class LeadAssignView(RoleRequiredMixin, View):
    """
    Lidni sotuvchiga biriktirish
    """
    allowed_roles = ['admin', 'manager', 'sales_manager']
    
    def post(self, request, pk):
        lead = get_object_or_404(Lead, pk=pk)
        sales_id = request.POST.get('sales_id')
        
        if sales_id:
            sales = get_object_or_404(User, pk=sales_id, role__in=['sales', 'sales_manager'])
            lead.assigned_sales = sales
            lead.assigned_at = timezone.now()
            lead.save()
            
            LeadHistory.objects.create(
                lead=lead,
                changed_by=request.user,
                notes=f"Sotuvchiga biriktirildi: {sales.username}"
            )
            
            messages.success(request, f'Lid {sales.username} ga biriktirildi.')
        
        return redirect('crm:lead_detail', pk=pk)


class LeadImportExcelView(RoleRequiredMixin, TemplateView):
    """
    Excel fayldan import
    """
    template_name = 'crm/excel_import.html'
    allowed_roles = ['admin', 'manager', 'sales_manager', 'sales']
    
    def post(self, request):
        file = request.FILES.get('excel_file')
        if not file:
            messages.error(request, 'Fayl tanlanmadi.')
            return redirect('crm:lead_list')
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            new_status = LeadStatus.objects.filter(code='new').first()
            imported = 0
            duplicates = 0
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0] or not row[1]:
                    continue
                
                name = str(row[0]).strip()
                phone = str(row[1]).strip()
                
                # Duplicate tekshirish
                if Lead.objects.filter(phone=phone).exists():
                    duplicates += 1
                    continue
                
                Lead.objects.create(
                    name=name,
                    phone=phone,
                    secondary_phone=str(row[2]).strip() if row[2] else None,
                    source='excel',
                    status=new_status,
                    created_by=request.user
                )
                imported += 1
            
            messages.success(request, f'{imported} ta lid import qilindi. {duplicates} ta dublikat.')
        
        except Exception as e:
            messages.error(request, f'Import xatosi: {str(e)}')
        
        return redirect('crm:lead_list')


class LeadGoogleSheetsImportView(RoleRequiredMixin, TemplateView):
    """
    Google Sheets dan import
    """
    template_name = 'crm/google_sheets_import.html'
    allowed_roles = ['admin', 'manager', 'sales_manager', 'sales']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from django.conf import settings
        context['sheets_configured'] = bool(getattr(settings, 'GOOGLE_SHEETS_ID', None))
        return context
    
    def post(self, request):
        from .tasks import import_leads_from_google_sheets
        import_leads_from_google_sheets.delay()
        messages.success(request, 'Google Sheets import boshlandi. Natijalar tez orada ko\'rinadi.')
        return redirect('crm:kanban')


# ==================== FOLLOW-UP VIEWS ====================

class FollowUpTodayView(LoginRequiredMixin, ListView):
    """
    Bugungi follow-up'lar
    """
    model = FollowUp
    template_name = 'crm/followup_today.html'
    context_object_name = 'followups'
    paginate_by = 25
    
    def get_queryset(self):
        queryset = FollowUp.objects.select_related('lead', 'sales')
        
        if self.request.user.is_sales:
            queryset = queryset.filter(sales=self.request.user)
        
        today = timezone.now().date()
        return queryset.filter(
            Q(due_date__date=today) | Q(due_date__lt=timezone.now(), completed=False)
        ).order_by('due_date')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['today_count'] = self.get_queryset().filter(due_date__date=timezone.now().date()).count()
        context['overdue_count'] = self.get_queryset().filter(is_overdue=True).count()
        return context


class FollowUpOverdueView(LoginRequiredMixin, ListView):
    """
    Kechikkan follow-up'lar
    """
    model = FollowUp
    template_name = 'crm/followup_overdue.html'
    context_object_name = 'followups'
    paginate_by = 25
    
    def get_queryset(self):
        queryset = FollowUp.objects.select_related('lead', 'sales').filter(
            completed=False,
            due_date__lt=timezone.now()
        )
        
        if self.request.user.is_sales:
            queryset = queryset.filter(sales=self.request.user)
        
        return queryset.order_by('due_date')


class FollowUpCompleteView(LoginRequiredMixin, View):
    """
    Follow-up ni bajarilgan deb belgilash
    """
    def post(self, request, pk):
        followup = get_object_or_404(FollowUp, pk=pk)
        
        # Ish vaqti tekshiruvi
        if request.user.is_sales:
            if hasattr(request.user, 'sales_profile') and request.user.sales_profile:
                profile = request.user.sales_profile
                if not profile.is_working_now():
                    messages.warning(request, 'Follow-up faqat ish vaqtida bajarilishi mumkin.')
                    return redirect('crm:followup_today')
        
        followup.completed = True
        followup.completed_at = timezone.now()
        followup.save()
        
        messages.success(request, 'Follow-up bajarildi.')
        
        next_url = request.POST.get('next', 'crm:followup_today')
        return redirect(next_url)


class FollowUpBulkRescheduleView(RoleRequiredMixin, View):
    """
    Bir nechta follow-up ni qayta rejalashtirish
    """
    allowed_roles = ['admin', 'manager', 'sales_manager']
    
    def post(self, request):
        followup_ids = request.POST.getlist('followup_ids')
        new_date = request.POST.get('new_date')
        
        if followup_ids and new_date:
            from datetime import datetime
            new_datetime = datetime.fromisoformat(new_date)
            
            FollowUp.objects.filter(id__in=followup_ids).update(
                due_date=new_datetime,
                is_overdue=False
            )
            
            messages.success(request, f'{len(followup_ids)} ta follow-up qayta rejalashtirildi.')
        
        return redirect('crm:followup_overdue')


class FollowUpBulkReassignView(RoleRequiredMixin, View):
    """
    Bir nechta follow-up ni boshqa sotuvchiga o'tkazish
    """
    allowed_roles = ['admin', 'manager', 'sales_manager']
    
    def post(self, request):
        followup_ids = request.POST.getlist('followup_ids')
        new_sales_id = request.POST.get('new_sales_id')
        
        if followup_ids and new_sales_id:
            new_sales = get_object_or_404(User, pk=new_sales_id)
            
            followups = FollowUp.objects.filter(id__in=followup_ids)
            for followup in followups:
                followup.sales = new_sales
                followup.lead.assigned_sales = new_sales
                followup.lead.save()
                followup.save()
            
            messages.success(request, f'{len(followup_ids)} ta follow-up {new_sales.username} ga o\'tkazildi.')
        
        return redirect('crm:followup_overdue')


class FollowUpBulkCompleteView(RoleRequiredMixin, View):
    """
    Bir nechta follow-up ni bajarish
    """
    allowed_roles = ['admin', 'manager', 'sales_manager']
    
    def post(self, request):
        followup_ids = request.POST.getlist('followup_ids')
        
        if followup_ids:
            FollowUp.objects.filter(id__in=followup_ids).update(
                completed=True,
                completed_at=timezone.now()
            )
            
            messages.success(request, f'{len(followup_ids)} ta follow-up bajarildi.')
        
        return redirect('crm:followup_overdue')


class FollowUpBulkDeleteView(AdminRequiredMixin, View):
    """
    Bir nechta follow-up ni o'chirish (faqat Admin)
    """
    def post(self, request):
        followup_ids = request.POST.getlist('followup_ids')
        
        if followup_ids:
            FollowUp.objects.filter(id__in=followup_ids).delete()
            messages.success(request, f'{len(followup_ids)} ta follow-up o\'chirildi.')
        
        return redirect('crm:followup_overdue')


class FollowUpListView(LoginRequiredMixin, ListView):
    """
    Follow-up'lar ro'yxati
    """
    model = FollowUp
    template_name = 'crm/followup_list.html'
    context_object_name = 'followups'
    paginate_by = 25
    
    def get_queryset(self):
        queryset = FollowUp.objects.select_related('lead', 'sales')
        
        if self.request.user.is_sales:
            queryset = queryset.filter(sales=self.request.user)
        
        filter_type = self.request.GET.get('filter', 'today')
        today = timezone.now().date()
        
        if filter_type == 'today':
            queryset = queryset.filter(due_date__date=today)
        elif filter_type == 'overdue':
            queryset = queryset.filter(is_overdue=True, completed=False)
        elif filter_type == 'upcoming':
            queryset = queryset.filter(due_date__gt=timezone.now(), completed=False)
        
        return queryset.order_by('due_date')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Permissions
        user = self.request.user
        context['can_create'] = user.is_superuser or (hasattr(user, 'is_admin') and user.is_admin) or (hasattr(user, 'is_manager') and user.is_manager) or (hasattr(user, 'is_sales_manager') and user.is_sales_manager) or (hasattr(user, 'is_sales') and user.is_sales)
        context['can_edit'] = context['can_create']
        context['can_delete'] = False  # Follow-ups usually shouldn't be deleted
        return context


class FollowUpCreateView(TailwindFormMixin, RoleRequiredMixin, CreateView):
    """
    Follow-up yaratish
    """
    model = FollowUp
    template_name = 'crm/followup_form.html'
    fields = ['lead', 'due_date', 'notes']
    allowed_roles = ['admin', 'manager', 'sales_manager', 'sales']
    success_url = reverse_lazy('crm:followup_today')
    
    def get_initial(self):
        initial = super().get_initial()
        lead_id = self.request.GET.get('lead')
        if lead_id:
            try:
                initial['lead'] = Lead.objects.get(pk=lead_id)
            except Lead.DoesNotExist:
                pass
        return initial
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        lead_id = self.request.GET.get('lead')
        if lead_id:
            try:
                context['lead'] = Lead.objects.get(pk=lead_id)
                context['back_url'] = 'crm:lead_detail'
                context['back_url_kwargs'] = lead_id
            except Lead.DoesNotExist:
                pass
        return context
    
    def form_valid(self, form):
        form.instance.sales = self.request.user
        messages.success(self.request, 'Follow-up muvaffaqiyatli yaratildi.')
        return super().form_valid(form)


class FollowUpUpdateView(TailwindFormMixin, RoleRequiredMixin, UpdateView):
    """
    Follow-up ni tahrirlash
    """
    model = FollowUp
    template_name = 'crm/followup_form.html'
    fields = ['due_date', 'notes', 'completed']
    allowed_roles = ['admin', 'manager', 'sales_manager', 'sales']
    success_url = reverse_lazy('crm:followup_today')
    
    def get_queryset(self):
        queryset = FollowUp.objects.all()
        if self.request.user.is_sales:
            queryset = queryset.filter(sales=self.request.user)
        return queryset
    
    def form_valid(self, form):
        if form.instance.completed and not form.instance.completed_at:
            form.instance.completed_at = timezone.now()
        messages.success(self.request, 'Follow-up muvaffaqiyatli yangilandi.')
        return super().form_valid(form)


# ==================== TRIAL LESSON VIEWS ====================

class TrialRegisterView(TailwindFormMixin, RoleRequiredMixin, CreateView):
    """
    Sinovga yozish
    """
    model = TrialLesson
    template_name = 'crm/trial_register.html'
    fields = ['group', 'room', 'date', 'time']
    allowed_roles = ['admin', 'manager', 'sales_manager', 'sales']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['lead'] = get_object_or_404(Lead, pk=self.kwargs['lead_pk'])
        context['groups'] = Group.objects.filter(is_active=True)
        context['rooms'] = Room.objects.filter(is_active=True)
        return context
    
    def form_valid(self, form):
        lead = get_object_or_404(Lead, pk=self.kwargs['lead_pk'])
        form.instance.lead = lead
        
        # Lead statusini yangilash
        trial_status = LeadStatus.objects.filter(code='trial_registered').first()
        if trial_status:
            lead.status = trial_status
            lead.trial_date = form.instance.date
            lead.trial_time = form.instance.time
            lead.trial_group = form.instance.group
            lead.trial_room = form.instance.room
            lead.save()
        
        messages.success(self.request, 'Sinov darsiga yozildi.')
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('crm:lead_detail', kwargs={'pk': self.kwargs['lead_pk']})


class TrialResultView(RoleRequiredMixin, View):
    """
    Sinov darsi natijasini kiritish
    """
    allowed_roles = ['admin', 'manager', 'sales_manager', 'sales']
    
    def post(self, request, pk):
        trial = get_object_or_404(TrialLesson, pk=pk)
        result = request.POST.get('result')
        notes = request.POST.get('notes', '')
        
        if result in ['attended', 'not_attended', 'accepted', 'rejected']:
            trial.result = result
            trial.notes = notes
            trial.save()
            
            # Lead statusini yangilash
            lead = trial.lead
            status_map = {
                'attended': 'trial_attended',
                'not_attended': 'trial_not_attended',
                'accepted': 'enrolled',
                'rejected': 'lost'
            }
            
            new_status = LeadStatus.objects.filter(code=status_map.get(result)).first()
            if new_status:
                lead.status = new_status
                lead.trial_result = result
                
                if result == 'accepted':
                    lead.enrolled_at = timezone.now()
                    lead.enrolled_group = trial.group
                elif result == 'rejected':
                    lead.lost_at = timezone.now()
                
                lead.save()
            
            messages.success(request, 'Sinov natijasi saqlandi.')
        
        return redirect('crm:lead_detail', pk=trial.lead.pk)


class TrialLessonCreateView(TailwindFormMixin, RoleRequiredMixin, CreateView):
    """
    Sinov darsiga yozish
    """
    model = TrialLesson
    template_name = 'crm/trial_lesson_form.html'
    fields = ['lead', 'group', 'room', 'date', 'time']
    allowed_roles = ['admin', 'manager', 'sales_manager', 'sales']
    success_url = reverse_lazy('crm:lead_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Sinov darsi muvaffaqiyatli yaratildi.')
        return super().form_valid(form)


class TrialLessonUpdateView(TailwindFormMixin, RoleRequiredMixin, UpdateView):
    """
    Sinov darsi natijasini kiritish
    """
    model = TrialLesson
    template_name = 'crm/trial_lesson_form.html'
    fields = ['result', 'notes']
    allowed_roles = ['admin', 'manager', 'sales_manager', 'sales']
    success_url = reverse_lazy('crm:lead_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Sinov darsi natijasi muvaffaqiyatli yangilandi.')
        return super().form_valid(form)


# ==================== OFFER VIEWS ====================

class OfferListView(LoginRequiredMixin, ListView):
    """
    Takliflar ro'yxati
    """
    model = Offer
    template_name = 'crm/offer_list.html'
    context_object_name = 'offers'
    paginate_by = 25
    
    def get_queryset(self):
        queryset = Offer.objects.select_related('course', 'created_by')
        
        offer_type = self.request.GET.get('type')
        if offer_type:
            queryset = queryset.filter(offer_type=offer_type)
        
        is_active = self.request.GET.get('active')
        if is_active == 'true':
            today = timezone.now().date()
            queryset = queryset.filter(is_active=True, valid_from__lte=today, valid_until__gte=today)
        
        return queryset.order_by('-created_at')


class OfferCreateView(TailwindFormMixin, RoleRequiredMixin, CreateView):
    """
    Taklif yaratish
    """
    model = Offer
    template_name = 'crm/offer_form.html'
    fields = ['title', 'description', 'offer_type', 'priority', 'channel', 'audience',
              'course', 'valid_from', 'valid_until', 'is_active']
    allowed_roles = ['admin', 'manager', 'sales_manager']
    success_url = reverse_lazy('crm:offer_list')
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, 'Taklif muvaffaqiyatli yaratildi.')
        return super().form_valid(form)


class OfferUpdateView(TailwindFormMixin, RoleRequiredMixin, UpdateView):
    """
    Taklifni tahrirlash
    """
    model = Offer
    template_name = 'crm/offer_form.html'
    fields = ['title', 'description', 'offer_type', 'priority', 'channel', 'audience',
              'course', 'valid_from', 'valid_until', 'is_active']
    allowed_roles = ['admin', 'manager', 'sales_manager']
    success_url = reverse_lazy('crm:offer_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Taklif muvaffaqiyatli yangilandi.')
        return super().form_valid(form)


class OfferDeleteView(AdminRequiredMixin, DeleteView):
    """
    Taklifni o'chirish
    """
    model = Offer
    template_name = 'crm/offer_confirm_delete.html'
    success_url = reverse_lazy('crm:offer_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Taklif o\'chirildi.')
        return super().delete(request, *args, **kwargs)


# ==================== LEAVE VIEWS ====================

class LeaveCreateView(TailwindFormMixin, RoleRequiredMixin, CreateView):
    """
    Ruxsat so'rash
    """
    model = Leave
    template_name = 'crm/leave_form.html'
    fields = ['start_date', 'end_date', 'start_time', 'end_time', 'reason']
    allowed_roles = ['sales', 'sales_manager']
    success_url = reverse_lazy('crm:leave_list')
    
    def form_valid(self, form):
        form.instance.sales = self.request.user
        messages.success(self.request, 'Ruxsat so\'rovi yuborildi.')
        return super().form_valid(form)


class LeaveListView(LoginRequiredMixin, ListView):
    """
    O'z ruxsatlari ro'yxati
    """
    model = Leave
    template_name = 'crm/leave_list.html'
    context_object_name = 'leaves'
    paginate_by = 25
    
    def get_queryset(self):
        if self.request.user.is_sales or self.request.user.is_sales_manager:
            return Leave.objects.filter(sales=self.request.user).order_by('-created_at')
        return Leave.objects.all().order_by('-created_at')


class LeavePendingView(RoleRequiredMixin, ListView):
    """
    Kutilayotgan ruxsat so'rovlari
    """
    model = Leave
    template_name = 'crm/leave_pending.html'
    context_object_name = 'leaves'
    allowed_roles = ['admin', 'manager', 'sales_manager']
    paginate_by = 25
    
    def get_queryset(self):
        return Leave.objects.filter(status='pending').select_related('sales').order_by('start_date')


class LeaveApproveView(RoleRequiredMixin, View):
    """
    Ruxsatni tasdiqlash/rad etish
    """
    allowed_roles = ['admin', 'manager', 'sales_manager']
    
    def post(self, request, pk):
        leave = get_object_or_404(Leave, pk=pk)
        action = request.POST.get('action')
        
        if action == 'approve':
            leave.status = 'approved'
            leave.approved_by = request.user
            leave.approved_at = timezone.now()
            leave.save()
            
            # SalesProfile yangilash
            if hasattr(leave.sales, 'sales_profile'):
                profile = leave.sales.sales_profile
                profile.is_on_leave = True
                profile.save()
            
            messages.success(request, 'Ruxsat tasdiqlandi.')
        
        elif action == 'reject':
            leave.status = 'rejected'
            leave.approved_by = request.user
            leave.rejection_reason = request.POST.get('rejection_reason', '')
            leave.save()
            messages.success(request, 'Ruxsat rad etildi.')
        
        return redirect('crm:leave_pending')


# ==================== SALES USER VIEWS ====================

class SalesUserListView(RoleRequiredMixin, ListView):
    """
    Sotuvchilar ro'yxati
    """
    model = SalesProfile
    template_name = 'crm/sales_list.html'
    context_object_name = 'sales_profiles'
    allowed_roles = ['admin', 'manager', 'sales_manager']
    paginate_by = 25
    
    def get_queryset(self):
        queryset = SalesProfile.objects.select_related('user', 'branch').filter(user__is_active=True)
        
        # Filterlar
        branch = self.request.GET.get('branch')
        if branch:
            queryset = queryset.filter(branch_id=branch)
        
        status = self.request.GET.get('status')
        if status == 'active':
            queryset = queryset.filter(is_active_sales=True, is_on_leave=False)
        elif status == 'on_leave':
            queryset = queryset.filter(is_on_leave=True)
        elif status == 'inactive':
            queryset = queryset.filter(is_active_sales=False)
        
        return queryset.order_by('user__first_name')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        profiles = self.get_queryset()
        
        # Basic counts
        context['total_count'] = profiles.count()
        context['active_count'] = profiles.filter(is_active_sales=True, is_on_leave=False).count()
        context['on_leave_count'] = profiles.filter(is_on_leave=True).count()
        
        # Get all sales users from filtered profiles
        sales_users = [profile.user for profile in profiles]
        
        # Overall statistics for filtered salespeople
        total_leads = Lead.objects.filter(assigned_sales__in=sales_users).count() if sales_users else 0
        total_enrolled = Lead.objects.filter(
            assigned_sales__in=sales_users,
            status__code='enrolled'
        ).count() if sales_users else 0
        avg_conversion = (total_enrolled / total_leads * 100) if total_leads > 0 else 0
        
        # Average KPI score (current month)
        current_month = timezone.now().month
        current_year = timezone.now().year
        kpis = SalesKPI.objects.filter(
            sales__in=sales_users,
            month=current_month,
            year=current_year
        ).aggregate(avg_kpi=Avg('total_kpi_score'))
        avg_kpi = kpis['avg_kpi'] or 0
        
        # Har bir sotuvchi uchun lid va sotuvlar soni
        for profile in context['sales_profiles']:
            profile.leads_count = Lead.objects.filter(assigned_sales=profile.user).count()
            profile.sales_count = Lead.objects.filter(
                assigned_sales=profile.user, 
                status__code='enrolled'
            ).count()
        
        # Overall statistics
        context['overall_stats'] = {
            'total_leads': total_leads,
            'total_enrolled': total_enrolled,
            'avg_conversion': avg_conversion,
            'avg_kpi': avg_kpi,
        }
        
        # Filter uchun filiallar
        context['branches'] = Branch.objects.all()
        
        return context


class SalesUserCreateView(RoleRequiredMixin, TemplateView):
    """
    Yangi sotuvchi yaratish
    """
    template_name = 'crm/sales_form.html'
    allowed_roles = ['admin', 'manager', 'sales_manager']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['branches'] = Branch.objects.all()
        context['is_edit'] = False
        return context
    
    def post(self, request):
        # User yaratish
        username = request.POST.get('username')
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Bu username allaqachon mavjud.')
            return redirect('crm:sales_create')
        
        user = User.objects.create(
            username=username,
            first_name=request.POST.get('first_name', ''),
            last_name=request.POST.get('last_name', ''),
            email=request.POST.get('email', ''),
            phone=request.POST.get('phone', ''),
            telegram_id=request.POST.get('telegram_chat_id') or None,
            role='sales',
        )
        password = request.POST.get('password', 'changeme123')
        user.set_password(password)
        user.save()
        
        # SalesProfile yaratish
        branch_id = request.POST.get('branch')
        SalesProfile.objects.create(
            user=user,
            branch_id=branch_id if branch_id else None,
            work_start_time=request.POST.get('work_start_time', '09:00'),
            work_end_time=request.POST.get('work_end_time', '18:00'),
            work_monday='work_monday' in request.POST,
            work_tuesday='work_tuesday' in request.POST,
            work_wednesday='work_wednesday' in request.POST,
            work_thursday='work_thursday' in request.POST,
            work_friday='work_friday' in request.POST,
            work_saturday='work_saturday' in request.POST,
            work_sunday='work_sunday' in request.POST,
        )
        
        messages.success(request, f'Sotuvchi muvaffaqiyatli yaratildi. Parol: {password}')
        return redirect('crm:sales_list')


class SalesUserUpdateView(RoleRequiredMixin, TemplateView):
    """
    Sotuvchini tahrirlash
    """
    template_name = 'crm/sales_form.html'
    allowed_roles = ['admin', 'manager', 'sales_manager']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        profile = get_object_or_404(SalesProfile, pk=self.kwargs['pk'])
        context['profile'] = profile
        context['user_obj'] = profile.user
        context['branches'] = Branch.objects.all()
        context['is_edit'] = True
        return context
    
    def post(self, request, pk):
        profile = get_object_or_404(SalesProfile, pk=pk)
        user = profile.user
        
        # User yangilash
        user.first_name = request.POST.get('first_name', '')
        user.last_name = request.POST.get('last_name', '')
        user.email = request.POST.get('email', '')
        user.phone = request.POST.get('phone', '')
        user.telegram_id = request.POST.get('telegram_chat_id') or None
        user.is_active = 'is_active' in request.POST
        user.save()
        
        # SalesProfile yangilash
        branch_id = request.POST.get('branch')
        profile.branch_id = branch_id if branch_id else None
        profile.work_start_time = request.POST.get('work_start_time', '09:00')
        profile.work_end_time = request.POST.get('work_end_time', '18:00')
        profile.work_monday = 'work_monday' in request.POST
        profile.work_tuesday = 'work_tuesday' in request.POST
        profile.work_wednesday = 'work_wednesday' in request.POST
        profile.work_thursday = 'work_thursday' in request.POST
        profile.work_friday = 'work_friday' in request.POST
        profile.work_saturday = 'work_saturday' in request.POST
        profile.work_sunday = 'work_sunday' in request.POST
        profile.is_active_sales = 'is_active_sales' in request.POST
        profile.save()
        
        messages.success(request, 'Sotuvchi muvaffaqiyatli yangilandi.')
        return redirect('crm:sales_list')


class SalesUserDeleteView(AdminRequiredMixin, DeleteView):
    """
    Sotuvchini o'chirish
    """
    model = User
    template_name = 'crm/sales_confirm_delete.html'
    success_url = reverse_lazy('crm:sales_list')
    
    def get_object(self):
        """
        PK can be either SalesProfile PK or User PK
        Handle both cases
        """
        pk = self.kwargs.get('pk')
        # Try to get SalesProfile first (since sales_list.html uses profile.pk)
        try:
            profile = SalesProfile.objects.get(pk=pk)
            return profile.user
        except SalesProfile.DoesNotExist:
            # If not SalesProfile, try User directly
            user = get_object_or_404(User, pk=pk, role='sales')
            return user
    
    def delete(self, request, *args, **kwargs):
        user = self.get_object()
        user_name = user.get_full_name() or user.username
        messages.success(request, f'Sotuvchi {user_name} o\'chirildi.')
        return super().delete(request, *args, **kwargs)


class SalesUserAbsenceView(RoleRequiredMixin, View):
    """
    Sotuvchini ishda emas deb belgilash
    """
    allowed_roles = ['admin', 'manager', 'sales_manager']
    
    def post(self, request, pk):
        user = get_object_or_404(User, pk=pk, role='sales')
        
        if hasattr(user, 'sales_profile'):
            profile = user.sales_profile
            profile.is_absent = request.POST.get('is_absent') == 'true'
            profile.absent_reason = request.POST.get('reason', '')
            profile.absent_from = timezone.now() if profile.is_absent else None
            profile.absent_until = request.POST.get('until') or None
            profile.save()
            
            messages.success(request, f'{user.username} ishda emaslik holati yangilandi.')
        
        return redirect('crm:sales_list')


# ==================== MANAGER USER VIEWS ====================

class ManagerUserListView(AdminRequiredMixin, ListView):
    """
    Menejerlar ro'yxati
    """
    model = User
    template_name = 'crm/manager_list.html'
    context_object_name = 'managers'
    paginate_by = 25
    
    def get_queryset(self):
        return User.objects.filter(role='sales_manager', is_active=True).order_by('username')


class ManagerUserCreateView(TailwindFormMixin, AdminRequiredMixin, CreateView):
    """
    Yangi menejer yaratish
    """
    model = User
    template_name = 'crm/manager_form.html'
    fields = ['username', 'first_name', 'last_name', 'email', 'phone']
    success_url = reverse_lazy('crm:manager_list')
    
    def form_valid(self, form):
        form.instance.role = 'sales_manager'
        form.instance.set_password('changeme123')
        response = super().form_valid(form)
        
        SalesProfile.objects.create(user=form.instance)
        
        messages.success(self.request, 'Menejer muvaffaqiyatli yaratildi. Default parol: changeme123')
        return response


class ManagerUserUpdateView(TailwindFormMixin, AdminRequiredMixin, UpdateView):
    """
    Menejerni tahrirlash
    """
    model = User
    template_name = 'crm/manager_form.html'
    fields = ['username', 'first_name', 'last_name', 'email', 'phone', 'is_active']
    success_url = reverse_lazy('crm:manager_list')
    
    def get_queryset(self):
        return User.objects.filter(role='sales_manager')


class ManagerUserDeleteView(AdminRequiredMixin, DeleteView):
    """
    Menejerni o'chirish
    """
    model = User
    template_name = 'crm/manager_confirm_delete.html'
    success_url = reverse_lazy('crm:manager_list')
    
    def get_queryset(self):
        return User.objects.filter(role='sales_manager')


# ==================== MESSAGE VIEWS ====================

class MessageCreateView(TailwindFormMixin, RoleRequiredMixin, CreateView):
    """
    Xabar yuborish
    """
    model = SalesMessage
    template_name = 'crm/message_form.html'
    fields = ['recipients', 'subject', 'message', 'priority']
    allowed_roles = ['admin', 'manager', 'sales_manager']
    success_url = reverse_lazy('crm:message_sent')
    
    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields['recipients'].queryset = User.objects.filter(
            role__in=['sales', 'sales_manager'], is_active=True
        )
        return form
    
    def form_valid(self, form):
        form.instance.sender = self.request.user
        messages.success(self.request, 'Xabar yuborildi.')
        return super().form_valid(form)


class MessageSentView(RoleRequiredMixin, ListView):
    """
    Yuborilgan xabarlar
    """
    model = SalesMessage
    template_name = 'crm/message_sent.html'
    context_object_name = 'messages_list'
    allowed_roles = ['admin', 'manager', 'sales_manager']
    paginate_by = 25
    
    def get_queryset(self):
        return SalesMessage.objects.filter(sender=self.request.user).order_by('-created_at')


class MessageInboxView(LoginRequiredMixin, ListView):
    """
    Kelgan xabarlar
    """
    model = SalesMessage
    template_name = 'crm/message_inbox.html'
    context_object_name = 'messages_list'
    paginate_by = 25
    
    def get_queryset(self):
        return SalesMessage.objects.filter(recipients=self.request.user).order_by('-created_at')


class MessageDetailView(LoginRequiredMixin, DetailView):
    """
    Xabar ko'rish
    """
    model = SalesMessage
    template_name = 'crm/message_detail.html'
    context_object_name = 'msg'
    
    def get(self, request, *args, **kwargs):
        response = super().get(request, *args, **kwargs)
        
        # O'qilgan deb belgilash
        if request.user in self.object.recipients.all():
            SalesMessageRead.objects.get_or_create(
                message=self.object,
                user=request.user
            )
        
        return response


class MessageDeleteView(RoleRequiredMixin, DeleteView):
    """
    Xabarni o'chirish
    """
    model = SalesMessage
    template_name = 'crm/message_confirm_delete.html'
    success_url = reverse_lazy('crm:message_sent')
    allowed_roles = ['admin', 'manager', 'sales_manager']
    
    def get_queryset(self):
        if self.request.user.is_admin:
            return SalesMessage.objects.all()
        return SalesMessage.objects.filter(sender=self.request.user)


# ==================== ANALYTICS VIEWS ====================

class CRMAnalyticsView(RoleRequiredMixin, TemplateView):
    """
    CRM statistika
    """
    template_name = 'crm/analytics.html'
    allowed_roles = ['admin', 'manager', 'sales_manager']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Umumiy statistika
        context['total_leads'] = Lead.objects.count()
        context['enrolled_leads'] = Lead.objects.filter(status__code='enrolled').count()
        context['lost_leads'] = Lead.objects.filter(status__code='lost').count()
        context['trial_leads'] = Lead.objects.filter(
            status__code__in=['trial_registered', 'trial_attended']
        ).count()
        
        # Bugungi statistika
        today = timezone.now().date()
        context['today_leads'] = Lead.objects.filter(created_at__date=today).count()
        context['today_followups'] = FollowUp.objects.filter(due_date__date=today).count()
        context['overdue_followups'] = FollowUp.objects.filter(
            completed=False, due_date__lt=timezone.now()
        ).count()
        
        # Konversiya
        total = Lead.objects.count()
        enrolled = Lead.objects.filter(status__code='enrolled').count()
        context['conversion_rate'] = (enrolled / total * 100) if total > 0 else 0
        
        # Sotuvchilar statistikasi
        context['sales_stats'] = User.objects.filter(
            role__in=['sales', 'sales_manager'], is_active=True
        ).annotate(
            lead_count=Count('assigned_leads'),
            enrolled_count=Count('assigned_leads', filter=Q(assigned_leads__status__code='enrolled'))
        ).order_by('-enrolled_count')[:10]
        
        # Manbalar bo'yicha
        context['source_stats'] = Lead.objects.values('source').annotate(
            count=Count('id')
        ).order_by('-count')
        
        return context


class SalesKPIListView(RoleRequiredMixin, TemplateView):
    """
    Barcha sotuvchilar KPI ko'rish (admin uchun)
    """
    template_name = 'crm/kpi_list.html'
    allowed_roles = ['admin', 'manager', 'sales_manager']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from django.db.models import Sum, Avg
        
        # Month/year filter (default to current month/year)
        month = int(self.request.GET.get('month', timezone.now().month))
        year = int(self.request.GET.get('year', timezone.now().year))
        
        context['selected_month'] = month
        context['selected_year'] = year
        
        # Get all sales users
        sales_users = User.objects.filter(role__in=['sales', 'sales_manager'], is_active=True)
        
        # Get KPI data for each salesperson for selected month/year
        sales_kpi_data = []
        for sales in sales_users:
            try:
                kpi = SalesKPI.objects.get(sales=sales, month=month, year=year)
            except SalesKPI.DoesNotExist:
                # Create default KPI if doesn't exist
                kpi = SalesKPI.objects.create(sales=sales, month=month, year=year)
            
            # Get additional stats
            total_leads = Lead.objects.filter(assigned_sales=sales).count()
            enrolled_leads = Lead.objects.filter(assigned_sales=sales, status__code='enrolled').count()
            conversion_rate = (enrolled_leads / total_leads * 100) if total_leads > 0 else 0
            
            sales_kpi_data.append({
                'sales': sales,
                'kpi': kpi,
                'total_leads': total_leads,
                'enrolled_leads': enrolled_leads,
                'conversion_rate': conversion_rate,
            })
        
        # Sort by KPI score
        sales_kpi_data.sort(key=lambda x: x['kpi'].total_kpi_score, reverse=True)
        context['sales_kpi_data'] = sales_kpi_data
        
        # Month options for dropdown
        context['months'] = [(i, timezone.datetime(2000, i, 1).strftime('%B')) for i in range(1, 13)]
        context['years'] = range(timezone.now().year - 2, timezone.now().year + 1)
        
        return context


class SalesKPIDetailView(RoleRequiredMixin, DetailView):
    """
    Sotuvchi KPI ko'rish
    """
    model = SalesKPI
    template_name = 'crm/sales_kpi_detail.html'
    context_object_name = 'kpi'
    allowed_roles = ['admin', 'manager', 'sales_manager', 'sales']
    
    def get_object(self):
        if self.request.user.is_sales:
            sales = self.request.user
        else:
            sales_id = self.kwargs.get('sales_id')
            if sales_id:
                sales = get_object_or_404(User, pk=sales_id, role__in=['sales', 'sales_manager'])
            else:
                sales = self.request.user
        
        # Get month/year from URL kwargs or query params
        month = self.kwargs.get('month') or int(self.request.GET.get('month', timezone.now().month))
        year = self.kwargs.get('year') or int(self.request.GET.get('year', timezone.now().year))
        
        kpi, created = SalesKPI.objects.get_or_create(
            sales=sales,
            month=month,
            year=year
        )
        
        return kpi


class SalesKPIRankingView(RoleRequiredMixin, ListView):
    """
    Sotuvchilar KPI reytingi
    """
    model = SalesKPI
    template_name = 'crm/sales_kpi_ranking.html'
    context_object_name = 'kpis'
    allowed_roles = ['admin', 'manager', 'sales_manager']
    paginate_by = 25
    
    def get_queryset(self):
        month = self.kwargs.get('month', timezone.now().month)
        year = self.kwargs.get('year', timezone.now().year)
        
        return SalesKPI.objects.filter(
            month=month,
            year=year
        ).select_related('sales').order_by('-total_kpi_score')


# ==================== LANDING VIEW ====================

class LandingView(TemplateView):
    """
    Landing sahifa (public)
    """
    template_name = 'crm/landing.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['courses'] = Course.objects.filter(is_active=True)[:6]
        return context


# Legacy views (backward compatibility)
class MessageListView(LoginRequiredMixin, ListView):
    """
    Xabarlar ro'yxati (legacy)
    """
    model = SalesMessage
    template_name = 'crm/message_list.html'
    context_object_name = 'messages_list'
    paginate_by = 25
    
    def get_queryset(self):
        if self.request.user.is_sales:
            return SalesMessage.objects.filter(recipients=self.request.user).order_by('-created_at')
        return SalesMessage.objects.filter(sender=self.request.user).order_by('-created_at')
